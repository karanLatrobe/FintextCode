from django.shortcuts import render, HttpResponseRedirect, redirect
from django.contrib import messages
from django.contrib.auth.forms import (
    AuthenticationForm, PasswordChangeForm, UserChangeForm, UserCreationForm
)
import threading
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import User
from django.views.decorators.cache import never_cache
from django.core.mail import send_mail
import re
from django.conf import settings
import pandas as pd
import os
from ml.fast_svm_predictor import fast_predict_latest_excel
import threading


def clean_text(x):
    x = str(x)
    x = re.sub(r'\b\d{1,2}/\d{1,2}/\d{2,4}\b', '', x)
    x = re.sub(r'\bValue\s*Date\b', '', x, flags=re.I)
    x = re.sub(r'\bValue\b', '', x, flags=re.I)
    x = re.sub(r'\bDate\b', '', x, flags=re.I)
    x = re.sub(r'Card\s+xx\d{4}', '', x, flags=re.I)
    x = re.sub(r'xx\d{4}', '', x)
    x = re.sub(r'\b\d{5,}\b', '', x)
    x = re.sub(r'[^a-zA-Z0-9\s/]', ' ', x)
    x = re.sub(r'\b\d+\b', '', x)
    x = re.sub(r'\s+', ' ', x).strip()
    return x
# ==============================================================
# HOME
# ==============================================================

def home(request):
    if request.user.is_authenticated:
        logout(request)
    return render(request, 'home.html')


# ==============================================================
# SIGNUP
# ==============================================================
import random
from django.contrib.auth import login
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import SignUP
from django.core.mail import send_mail

def sign_up(request):
    if request.method == 'POST':
        form = SignUP(request.POST)
        if form.is_valid():
            request.session['temp_user'] = form.cleaned_data
            otp = random.randint(100000, 999999)

            request.session['otp'] = otp

            send_mail(
                'Your OTP Verification',
                f'Your OTP is: {otp}',
                'yourgmail@gmail.com',
                [form.cleaned_data['email']],
                fail_silently=False,
            )

            return redirect('verify_otp')

    else:
        form = SignUP()

    return render(request, 'signup.html', {'form': form})


from django.contrib.auth.models import User

def verify_otp(request):

    # ⛔ Direct URL Access Ko Block Karna
    if 'otp' not in request.session or 'temp_user' not in request.session:
        return redirect('signup')

    if request.method == "POST":
        entered = request.POST.get('otp')
        real_otp = str(request.session.get('otp'))
        user_data = request.session.get('temp_user')

        if entered == real_otp:
            user = User.objects.create_user(
                username=user_data['username'],
                email=user_data['email'],
                password=user_data['password1']
            )
            user.save()

            messages.success(request, "Account Verified Successfully!")
            
            del request.session['otp']
            del request.session['temp_user']

            return redirect('login')
        else:
            messages.error(request, "Invalid OTP. Try again.")

    return render(request, 'verify_otp.html')


# ==============================================================
# LOGIN
# ==============================================================


@never_cache
def user_login(request):

    # ❌ Logged-in user ko logout mat karo
    # ✔ Instead usse profile pe redirect karo
    if request.user.is_authenticated:
        return redirect('profile')

    if request.method == 'POST':
        form = AuthenticationForm(request=request, data=request.POST)

        if form.is_valid():
            uname = form.cleaned_data['username']
            upass = form.cleaned_data['password']
            user = authenticate(username=uname, password=upass)

            if user is not None:
                login(request, user)

                # Login success email
                try:
                    send_mail(
                        subject='Login Notification',
                        message=f'User {uname} logged in successfully.',
                        from_email='karansharma89200@gmail.com',
                        recipient_list=['myquoraai89200@gmail.com'],
                        fail_silently=True,
                    )
                except:
                    pass

                messages.success(request, f"Hi {request.user.username}, Logged in successfully!")
                return redirect('profile')

        else:
            messages.error(request, "Invalid username or password.")

    else:
        form = AuthenticationForm()

    return render(request, 'userlogin.html', {'form': form})




def forgot_password(request):
    if request.method == "POST":
        email = request.POST.get("email")

        # Check user exists
        if not User.objects.filter(email=email).exists():
            messages.error(request, "Email not registered!")
            return redirect("forgot_password")

        # OTP generate
        otp = random.randint(100000, 999999)

        request.session['reset_email'] = email
        request.session['reset_otp'] = otp

        # Send OTP
        send_mail(
            "Password Reset OTP",
            f"Your OTP is: {otp}",
            "karansharma89200@gmail.com",
            [email],
            fail_silently=False
        )

        return redirect("forgot_password_otp")

    return render(request, "forgot_password.html")


def forgot_password_otp(request):

    # Direct URL access block
    if 'reset_email' not in request.session:
        return redirect("forgot_password")

    if request.method == "POST":
        entered = request.POST.get("otp")
        real = str(request.session.get("reset_otp"))

        if entered == real:
            return redirect("reset_password")
        else:
            messages.error(request, "Invalid OTP!")

    return render(request, "forgot_password_otp.html")


def reset_password(request):

    # Direct URL access block
    if 'reset_email' not in request.session:
        return redirect("forgot_password")

    if request.method == "POST":
        new = request.POST.get("new_password")
        confirm = request.POST.get("confirm_password")

        if new != confirm:
            messages.error(request, "Passwords do not match!")
            return redirect("reset_password")

        email = request.session['reset_email']
        user = User.objects.get(email=email)
        user.set_password(new)
        user.save()

        del request.session['reset_email']
        del request.session['reset_otp']

        messages.success(request, "Password updated! Login now.")
        return redirect("login")

    return render(request, "reset_password.html")


# ==============================================================
# LOGOUT
# ==============================================================

def user_logout(request):
    logout(request)
    return redirect('home')


# ==============================================================
# CHANGE PASSWORD
# ==============================================================

def user_change_pass(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            PS = PasswordChangeForm(user=request.user, data=request.POST)
            if PS.is_valid():
                PS.save()
                update_session_auth_hash(request, PS.user)
                return HttpResponseRedirect('/profile/')
        else:
            PS = PasswordChangeForm(user=request.user)

        return render(request, 'changepass.html', {'form': PS})

    return HttpResponseRedirect('/login/')


# ==============================================================
# USER DETAIL (ADMIN VIEW)
# ==============================================================

def user_detail(request, id):
    if request.user.is_authenticated:
        pi = User.objects.get(pk=id)
        fm = EditAdminProfileForm(instance=pi)
        return render(request, "user_detail.html", {'form': fm})
    return HttpResponseRedirect('/login/')


# ==============================================================
# IMPORTS FOR PROFILE + PDF/CSV PROCESSING
# ==============================================================

import os
import datetime
from pathlib import Path
import pandas as pd

from django.http import HttpResponse, Http404, JsonResponse
from django.conf import settings
from django.urls import reverse

from enroll.forms import EditUserProfileForm, EditAdminProfileForm
from enroll.utils.common import process_pdf_and_return_output_path


# ==============================================================
# MAIN PROFILE VIEW
# ==============================================================
# ==============================================================
# MAIN PROFILE VIEW
# ==============================================================

@never_cache
def user_profile(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect('/login/')

    name = request.user.username
    users = User.objects.all() if request.user.is_superuser else None

    rows = None
    columns = None
    download_link = None

    # ------------------ CASE 1: FILE UPLOAD (PDF/CSV) ------------------
    if "uploadedFile" in request.FILES:
        pdf = request.FILES["uploadedFile"]

        temp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
        os.makedirs(temp_dir, exist_ok=True)

        temp_pdf = os.path.join(temp_dir, pdf.name)
        with open(temp_pdf, "wb+") as f:
            for c in pdf.chunks():
                f.write(c)

        try:
            raw_excel_path, df = process_pdf_and_return_output_path(temp_pdf)
            print(df.columns)
            for col in ["Debit", "Credit", "Balance"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce")

            df = df.fillna("")
        except Exception as e:
            return render(request, "profile.html", {"name": name, "msg": f"PDF Processing Error: {e}"})

        result_df = df.copy()

        outputs_dir = Path(__file__).resolve().parent / "utils" / "outputs"
        outputs_dir.mkdir(exist_ok=True, parents=True)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"statement_{timestamp}.xlsx"
        out_full_path = outputs_dir / out_name
        result_df.to_excel(out_full_path, index=False)

        # =============== UPDATE TRAINING FILE FIRST TIME ====================
        training_dir = os.path.join(settings.BASE_DIR, "TrainingFile")
        os.makedirs(training_dir, exist_ok=True)

        training_excel = os.path.join(training_dir, "LatestTraining.xlsx")

        # Always overwrite with first extracted file
        result_df.to_excel(training_excel, index=False)
        # ================================================================

        rows = result_df.to_dict(orient="records")
        columns = result_df.columns.tolist()
        download_link = f"{reverse('download_excel')}?file={out_name}"

        os.remove(temp_pdf)

        return render(request, "profile.html", {
            "name": name,
            "rows": rows,
            "columns": columns,
            "download_link": download_link,
            "users": users
        })

    # ------------------ CASE 2: PROFILE UPDATE ------------------
    if request.method == "POST":
        if request.user.is_superuser:
            fm = EditAdminProfileForm(request.POST, instance=request.user)
        else:
            fm = EditUserProfileForm(request.POST, instance=request.user)

        if fm.is_valid():
            fm.save()
            messages.success(request, "Profile Updated!")
    else:
        if request.user.is_superuser:
            fm = EditAdminProfileForm(instance=request.user)
        else:
            fm = EditUserProfileForm(instance=request.user)

    return render(request, "profile.html", {
        "name": name,
        "form": fm,
        "users": users,
        "rows": rows,
        "columns": columns,
        "download_link": download_link,
    })


# ==============================================================
# DOWNLOAD EXCEL
# ==============================================================

from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from django.http import HttpResponse, Http404
from pathlib import Path

def download_excel(request):
    fname = request.GET.get("file")

    if not fname:
        raise Http404("File missing")

    outputs_dir = Path(__file__).resolve().parent / "utils" / "outputs"
    file_path = outputs_dir / fname

    if not file_path.exists():
        raise Http404("Not found")

    # ---------- READ DATA ----------
    df = pd.read_excel(file_path)

    target_cols = ["Debit", "Credit", "Balance"]

    for col in target_cols:
        if col in df.columns:
            df[col] = (
                df[col]
                .replace("-", pd.NA)
                .astype(str)
                .str.replace(",", "", regex=False)
            )
            df[col] = pd.to_numeric(df[col], errors="coerce").round(4)

    # ---------- CREATE EXCEL MANUALLY ----------
    wb = Workbook()
    ws = wb.active

    # Write headers
    ws.append(list(df.columns))

    # 🔥 INSERT ONE NUMERIC ROW (Excel type trigger)
    temp_row = []
    for col in df.columns:
        if col in target_cols:
            temp_row.append(0.0000)
        else:
            temp_row.append("")
    ws.append(temp_row)

    # Write actual data
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Remove temp row
    ws.delete_rows(2)

    # Force number format
    for col in target_cols:
        if col in df.columns:
            idx = df.columns.get_loc(col) + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = "0.0000"

    # ---------- RETURN FILE ----------
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.ms-excel"
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response

# ==============================================================
# AJAX UPLOAD
# ==============================================================

from enroll.models import ExcelRow
import uuid

from enroll.models import ExcelRow
import uuid

import tempfile

def ajax_upload(request):
    if request.method == "POST" and request.FILES.get("uploadedFile"):

        uploaded = request.FILES["uploadedFile"]

        # ----------------------------
        # 1) Save uploaded file temp location
        # ----------------------------
        import tempfile
        temp = tempfile.NamedTemporaryFile(delete=False, suffix=uploaded.name)
        for chunk in uploaded.chunks():
            temp.write(chunk)
        temp_path = temp.name
        temp.close()

        # ----------------------------
        # 2) Convert file → DataFrame
        # ----------------------------
        try:
            if uploaded.name.lower().endswith(".pdf"):
                raw_excel_path, df = process_pdf_and_return_output_path(temp_path)
            else:
                df = pd.read_csv(temp_path, on_bad_lines="skip", encoding_errors="ignore")

            for col in ["Debit", "Credit", "Balance"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce")

            df = df.fillna("")



            # ============================================================
            # 🔥 FIRST TIME TRAINING FILE UPDATE (MAIN FIX)
            # ============================================================
            from django.conf import settings
            training_dir = os.path.join(settings.BASE_DIR, "TrainingFile")
            os.makedirs(training_dir, exist_ok=True)

            training_excel = os.path.join(training_dir, "LatestTraining.xlsx")

            # overwrite Training File immediately after extraction
            df.to_excel(training_excel, index=False)
            # ================= RUN ML PREDICTION =================

            fast_predict_latest_excel()


            # Now read predicted excel
            df = pd.read_excel(training_excel).fillna("-")
# ====================================================

            # ============================================================

        except Exception as e:
            return JsonResponse({"error": f"Processing error: {str(e)}"})

        result_df = df.copy()

# 🔥 FIX: JSONField-safe conversion
        safe_rows = []

        # ----------------------------
        # 3) Create unique file ID
        # ----------------------------
        import uuid
        file_id = str(uuid.uuid4())

        # ----------------------------
        # 4) Save rows into DB
        # ----------------------------
        for i, row in result_df.iterrows():
            safe_rows.append(
                ExcelRow(
                    user=request.user,
                    file_id=file_id,
                    row_index=i,
                    data={
                        k: (v.isoformat() if hasattr(v, "isoformat") else v)
                        for k, v in row.to_dict().items()
                    }
                )
            )
        ExcelRow.objects.bulk_create(safe_rows)
        # ----------------------------
        # 5) Return JSON Response
        # ----------------------------
        return JsonResponse({
            "columns": list(result_df.columns),
            "rows": result_df.to_dict(orient="records"),
            "file_id": file_id
        })

    return JsonResponse({"error": "Invalid request"}, status=400)

# ==============================================================
# FAQ
# ==============================================================

# ==============================================================  
# SAVE UPDATED EXCEL (FINAL STABLE VERSION)  
# ==============================================================  

import pandas as pd
from django.http import JsonResponse
from pathlib import Path
import json

def save_updated_excel(request):
    """
    ✔ Frontend se updated rows + columns + filename receive karta hai
    ✔ Editable duplicate Excel file overwrite karta hai
    ✔ Download ke time latest updated file hi milegi
    """
    if request.method == "POST":
        try:
            # JSON data read
            data = json.loads(request.body)

            filename = data.get("file")
            columns = data.get("columns")
            rows = data.get("rows")

            if not filename:
                return JsonResponse({"success": False, "error": "Filename missing"})

            # Editable duplicate file ka path
            outputs_dir = Path(__file__).resolve().parent / "utils" / "outputs"
            file_path = outputs_dir / filename

            # Check file exists
            if not file_path.exists():
                return JsonResponse({"success": False, "error": "Editable file not found"})

            # JSON rows → DataFrame
            df = pd.DataFrame(rows, columns=columns)

            # Overwrite editable Excel
            df.to_excel(file_path, index=False)

            return JsonResponse({"success": True})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid Request"})


def faq(request):
    return render(request, 'FAQ.html')


from enroll.models import ExcelRow
import json


import os
import pandas as pd
import json
from django.http import JsonResponse
from enroll.models import ExcelRow

def save_updated_db(request):
    if request.method == "POST":
        data = json.loads(request.body)
        file_id = data["file_id"]
        rows = data["rows"]

        # ================================
        # STEP 0: CSV PATH (MISSING BUG FIX)
        # ================================
        TRAINING_CSV = os.path.join(
            settings.BASE_DIR,
            "TrainingFile",
            "TrainingFile.csv"
        )

        # ================================
        # STEP 1: FETCH OLD ROWS FIRST ❗
        # ================================
        old_rows_qs = ExcelRow.objects.filter(
            file_id=file_id,
            user=request.user
        ).order_by("row_index")

        old_df = pd.DataFrame([r.data for r in old_rows_qs])

        # ================================
        # STEP 2: NEW DATAFRAME
        # ================================
        new_df = pd.DataFrame(rows)

        # ================================
        # STEP 3: FIND CHANGED ROWS
        # ================================
        rows_to_add = []

        for i in range(min(len(old_df), len(new_df))):

            old_txn = str(old_df.iloc[i, 1])   # 🔥 2nd column
            new_txn = str(new_df.iloc[i, 1])

            old_cat = str(old_df.iloc[i].get("Account Type", ""))
            new_cat = str(new_df.iloc[i].get("Account Type", ""))

            if old_txn != new_txn or old_cat != new_cat:
                rows_to_add.append({
                    "Transaction_clean": new_txn,
                    "Account": new_cat
                })

        # ================================
        # STEP 4: APPEND TO TrainingFile.csv
        # ================================
        if rows_to_add:
            os.makedirs(os.path.dirname(TRAINING_CSV), exist_ok=True)

            write_header = not os.path.exists(TRAINING_CSV)

            pd.DataFrame(rows_to_add).to_csv(
                TRAINING_CSV,
                mode="a",
                index=False,
                header=write_header
            )

        # ================================
        # STEP 5: NOW DELETE & SAVE DB
        # ================================
        ExcelRow.objects.filter(
            file_id=file_id,
            user=request.user
        ).delete()

        ExcelRow.objects.bulk_create([
            ExcelRow(
                user=request.user,
                file_id=file_id,
                row_index=i,
                data=row
            )
            for i, row in enumerate(rows)
        ])

        # ================================
        # STEP 6: UPDATE LatestTraining.xlsx
        # ================================
        training_dir = os.path.join(settings.BASE_DIR, "TrainingFile")
        os.makedirs(training_dir, exist_ok=True)

        training_excel = os.path.join(training_dir, "LatestTraining.xlsx")
        new_df.to_excel(training_excel, index=False)

        # ================= RUN ML AGAIN =================
        threading.Thread(
            target=fast_predict_latest_excel,
            daemon=True
        ).start()

        return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Invalid request"})


import os
import tempfile

def download_excel_from_db(request):
    file_id = request.GET.get("file_id")

    rows = ExcelRow.objects.filter(
        file_id=file_id,
        
    ).order_by("row_index")

    if not rows.exists():
        return HttpResponse("No data found", status=404)

    df = pd.DataFrame([r.data for r in rows])

    for col in ["Debit", "Credit", "Balance"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, f"{file_id}.xlsx")

    df.to_excel(temp_path, index=False)

    with open(temp_path, "rb") as f:
        data = f.read()

    response = HttpResponse(data, content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = "attachment; filename=updated_data.xlsx"
    return response
